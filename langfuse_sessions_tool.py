import io
import time
import requests
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta

from ibm_watsonx_orchestrate.agent_builder.tools import tool
from ibm_watsonx_orchestrate.agent_builder.connections import ConnectionType
from ibm_watsonx_orchestrate.run import connections

LANGFUSE_APP_ID = "m-langfuse"
JST = timezone(timedelta(hours=9))
# サービス本番開始: JST 2026/05/01 00:00 = UTC 2026/04/30 15:00
SERVICE_START_UTC = datetime(2026, 4, 30, 15, 0, 0, tzinfo=timezone.utc)

HEADER_BG = "375623"
EVEN_ROW_BG = "EBF5E0"
ROW_HEIGHT = 15
FONT_NAME = "游ゴシック"

COLUMNS = [
    ("sessionId",             "セッションID",         28),
    ("start_time",            "開始時刻 (JST)",       18),
    ("turn_count",            "ターン数",              10),
    ("total_latency",         "合計レイテンシ(秒)",   18),
    ("userId",                "ユーザーID",            28),
    ("agent_id",              "エージェントID",        36),
    ("agent_display_name",    "エージェント名",        20),
    ("first_user_message",    "最初のユーザー発言",   42),
    ("last_assistant_message","最後のAI応答",          42),
]

RIGHT_ALIGNED = {"turn_count", "total_latency"}


@tool(
    name="m_export_langfuse_sessions",
    display_name="M - Langfuse セッション集計 Excel エクスポート",
    expected_credentials=[{"app_id": LANGFUSE_APP_ID, "type": ConnectionType.KEY_VALUE}],
)
def export_langfuse_sessions() -> bytes:
    """Langfuse からトレースを取得し、セッション単位に集計した Excel ファイルをバイト列で返す。

    Returns:
        bytes: セッション集計 Excel (.xlsx) のバイト列（ファイルダウンロード用）。
    """
    creds = connections.key_value(LANGFUSE_APP_ID)
    public_key = creds.get("LANGFUSE_PUBLIC_KEY")
    secret_key = creds.get("LANGFUSE_SECRET_KEY")
    host = creds.get("LANGFUSE_HOST", "https://cloud.langfuse.com")

    from_ts = SERVICE_START_UTC.isoformat()
    traces = _fetch_all_traces(public_key, secret_key, host, from_ts)
    obs_by_trace = _fetch_all_observations(public_key, secret_key, host, from_ts)
    rows = _aggregate_sessions(traces, obs_by_trace)
    return _build_xlsx(rows)


# ---------- Langfuse API ----------

def _fetch_page(public_key: str, secret_key: str, host: str, endpoint: str, params: dict) -> list:
    for attempt in range(4):
        resp = requests.get(
            f"{host}/api/public/{endpoint}",
            auth=(public_key, secret_key),
            params=params,
            timeout=30,
        )
        if resp.status_code == 429:
            time.sleep(0.5 * (2 ** attempt))
            continue
        resp.raise_for_status()
        return resp.json().get("data", [])
    return []


def _fetch_all_pages(public_key: str, secret_key: str, host: str, endpoint: str, base_params: dict) -> list:
    p1 = requests.get(
        f"{host}/api/public/{endpoint}",
        auth=(public_key, secret_key),
        params={**base_params, "page": 1},
        timeout=30,
    )
    p1.raise_for_status()
    body = p1.json()
    items = body.get("data", [])
    meta = body.get("meta", {})
    total_pages = meta.get("totalPages") or max(
        1, (meta.get("totalItems", 0) + base_params["limit"] - 1) // base_params["limit"]
    )

    if total_pages <= 1:
        return items

    results = list(items)
    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {
            ex.submit(_fetch_page, public_key, secret_key, host, endpoint, {**base_params, "page": p}): p
            for p in range(2, total_pages + 1)
        }
        for fut in as_completed(futs):
            results.extend(fut.result())
    return results


def _fetch_all_traces(public_key: str, secret_key: str, host: str, from_timestamp: str) -> list:
    return _fetch_all_pages(
        public_key, secret_key, host,
        "traces",
        {"limit": 50, "fromTimestamp": from_timestamp},
    )


def _fetch_all_observations(public_key: str, secret_key: str, host: str, from_timestamp: str) -> dict:
    """LangGraph observations を一括取得して {traceId: [obs, ...]} で返す。"""
    items = _fetch_all_pages(
        public_key, secret_key, host,
        "observations",
        {"limit": 50, "fromTimestamp": from_timestamp, "name": "LangGraph"},
    )
    obs_by_trace: dict = defaultdict(list)
    for obs in items:
        tid = obs.get("traceId", "")
        if tid:
            obs_by_trace[tid].append(obs)
    return obs_by_trace


def _get_user_message(trace: dict, obs_list: list | None = None) -> str:
    try:
        inp = trace.get("input") or {}
        for m in inp.get("messages", []):
            if m.get("role") == "user":
                content = m.get("content", "")
                return (content if isinstance(content, str) else str(content))[:200]
    except Exception:
        pass
    if obs_list:
        for obs in obs_list:
            try:
                inp = obs.get("input") or {}
                if isinstance(inp, dict):
                    for m in inp.get("messages", []):
                        if m.get("role") == "user":
                            content = m.get("content", "")
                            return (content if isinstance(content, str) else str(content))[:200]
            except Exception:
                pass
    return ""


def _get_assistant_message(trace: dict, obs_list: list | None = None) -> str:
    try:
        out = trace.get("output") or {}
        for m in out.get("messages", []):
            if m.get("role") == "assistant":
                content = m.get("content", "")
                return (content if isinstance(content, str) else str(content))[:200]
    except Exception:
        pass
    if obs_list:
        for obs in reversed(obs_list or []):
            try:
                out = obs.get("output") or {}
                if isinstance(out, dict):
                    for m in out.get("messages", []):
                        if m.get("role") == "assistant":
                            content = m.get("content", "")
                            return (content if isinstance(content, str) else str(content))[:200]
            except Exception:
                pass
    return ""


def _get_agent_info(trace: dict, obs_list: list | None = None) -> tuple:
    try:
        inp = trace.get("input") or {}
        agent_id = inp.get("current_agent_id", "")
        agent_name = inp.get("agent_display_name", "")
        if agent_id or agent_name:
            return agent_id, agent_name
    except Exception:
        pass
    if obs_list:
        for obs in obs_list:
            try:
                inp = obs.get("input") or {}
                if isinstance(inp, dict):
                    agent_id = inp.get("current_agent_id", "")
                    agent_name = inp.get("agent_display_name", "")
                    if agent_id or agent_name:
                        return agent_id, agent_name
            except Exception:
                pass
    return "", ""


def _to_jst(ts_str: str) -> str:
    if not ts_str:
        return ""
    try:
        dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        return dt.astimezone(JST).strftime("%Y/%m/%d %H:%M")
    except Exception:
        return ts_str


# ---------- 集計 ----------

def _aggregate_sessions(traces: list, obs_by_trace: dict) -> list:
    sessions: dict = defaultdict(lambda: {
        "start_time": None,
        "turn_count": 0,
        "total_latency": 0.0,
        "userId": "",
        "agent_id": "",
        "agent_display_name": "",
        "first_user_message": "",
        "last_assistant_message": "",
    })

    for t in sorted(traces, key=lambda x: x.get("timestamp", "")):
        sid = t.get("sessionId", "unknown")
        ts = t.get("timestamp", "")
        s = sessions[sid]
        obs_list = obs_by_trace.get(t.get("id", ""), [])

        if s["start_time"] is None or ts < s["start_time"]:
            s["start_time"] = ts

        s["turn_count"] += 1
        s["total_latency"] += t.get("latency") or 0.0
        s["userId"] = s["userId"] or t.get("userId", "")

        agent_id, agent_name = _get_agent_info(t, obs_list)
        s["agent_id"] = s["agent_id"] or agent_id
        s["agent_display_name"] = s["agent_display_name"] or agent_name

        user_msg = _get_user_message(t, obs_list)
        if user_msg and not s["first_user_message"]:
            s["first_user_message"] = user_msg

        asst_msg = _get_assistant_message(t, obs_list)
        if asst_msg:
            s["last_assistant_message"] = asst_msg

    rows = []
    for sid, s in sorted(sessions.items(), key=lambda x: x[1]["start_time"] or "", reverse=True):
        rows.append({
            "sessionId":              sid,
            "start_time":             _to_jst(s["start_time"]),
            "turn_count":             s["turn_count"],
            "total_latency":          round(s["total_latency"], 3),
            "userId":                 s["userId"],
            "agent_id":               s["agent_id"],
            "agent_display_name":     s["agent_display_name"],
            "first_user_message":     s["first_user_message"],
            "last_assistant_message": s["last_assistant_message"],
        })
    return rows


# ---------- Excel 出力 ----------

def _build_xlsx(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "セッション集計"

    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    even_fill   = PatternFill("solid", fgColor=EVEN_ROW_BG)
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    data_font   = Font(name=FONT_NAME)

    # ヘッダー行
    for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = ROW_HEIGHT

    # データ行
    for row_idx, row in enumerate(rows, start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, (key, _label, _width) in enumerate(COLUMNS, start=1):
            value = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.font = data_font
            align = "right" if key in RIGHT_ALIGNED else "left"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
